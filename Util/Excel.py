#处理excel的一些操作
from openpyxl.styles import Font, colors, NamedStyle, Side, Border, PatternFill
from openpyxl import *
from Util.DateAndTime import *
import os




class ExcelUtil():

    def __init__(self,file_path):
        self.file_path = file_path
        if not os.path.exists(file_path) or not (".xlsx" in file_path):
            print("指定的 excel 文件路径%s不存在,或者文件类型不是xlsx"%file_path)
        else:
            self.wb = load_workbook(file_path)
            print("加载 excel %s 文件成功"%file_path)

        self.sheet = self.wb.active #默认打开的sheet是谁，也就是上一次默认打开的是哪一个sheet

    def get_sheet_names(self):#获取所有sheet名称
        return self.wb.sheetnames

    def set_sheet_names_by_index(self,index):#通过index指定操作哪个sheet，设置默认要操作哪个sheet
        if not isinstance(index,int):#判断参数是否是整数
            print("您设定的sheet序号 '%s' 不是整数,请重新设定"%index)
            return
        if not index <=len(self.get_sheet_names()):#判断index是否大于0，并且小于sheet的最大数量
            print("您设定的sheet序号 '%s' 不存在,请重新设定"%index)
            return
        else:#设定好第几个sheet
            self.sheet = self.wb[self.get_sheet_names()[index - 1]]
        return self.sheet

    def set_sheet_by_name(self,sheet_name):#通过name指定操作哪个sheet
        if not sheet_name in self.get_sheet_names():
            print("设定的sheet名称：%s 不存在，请重新设定！"%sheet_name)
            return

        self.sheet = self.wb[sheet_name]
        return self.sheet

    #读取某一个单元格，读取某一行，读取某一列
    def get_sheet_all_data(self):#读取所有单元格的数据
        data = []
        for row in self.sheet.iter_rows():#遍历所有的行
            row_data = []
            for cell in row:#遍历一行的所有单元格
                row_data.append(cell.value)
            data.append(row_data)
        return data



    def get_max_row_no(self):#最大行号
        return self.sheet.max_row


    def get_max_col_no(self):#最大列号
        return self.sheet.max_column


    def write_lines_in_sheet(self,data):#向单元格写入多行数据
        if not isinstance(data,(list,tuple)):
            print("您写入的数据不是列表或者元组类型，请重新设定")
        for line in data:
            if not isinstance(line,(list,tuple)):
                print("您写入的数据不是列表或者元组类型，请重新设定")
                return
            self.sheet.append(line)

        bd = Side(style = 'thin',color = "000000")
        for row in self.sheet.rows:
            for cell in row:
                cell.border = Border(left = bd,top=bd,right = bd,bottom = bd)

        self.wb.save(self.file_path)
        print("excel 数据已经全部写入！")

    def write_a_line_in_sheet(self,data,font_color = None,fgcolor = None):#向单元格写一行数据
        if fgcolor is not None:
            fill = PatternFill(fill_type="solid",fgColor = fgcolor)
        else:
            fill = None

        if font_color is None:
            ft = None
        elif "red" in font_color:
            ft = Font(color = colors.RED)
        elif "green" in font_color:
            ft = Font(color="00FF00")

        if not isinstance(data,(list,tuple)):
            print("您写入的数据不是列表或者元组类型，请重新设定")
            return
        rowNo = self.get_max_row_no() + 1
        for idx,value in enumerate(data):
            if fill is not None:
                self.sheet.cell(row = rowNo,column=idx+1).fill = fill
            if ((data[idx] == "成功") or \
                    (data[idx] == "失败")) and ft is not None:
                self.sheet.cell(row=rowNo, column=idx + 1).font = ft
            self.sheet.cell(row=rowNo, column=idx + 1).value = data[idx]
        #self.sheet.append(data)

        bd = Side(style = 'thin',color = "000000")
        for row in self.sheet.rows:
            for cell in row:
                cell.border = Border(left = bd,top=bd,right = bd,bottom = bd)




    def write_col_in_sheet(self,col_no,data):#向单元格写入多列数据
        if not isinstance(data,(list,tuple)):
            print("您写入的数据不是列表或者元组类型，请重新设定")
            return None
        num = len(data)
        for i in range(num):
            self.write_cell_value(i,col_no,data[i])

        self.wb.save(self.file_path)
        print("excel 数据已经全部写入！")


    def create_new_sheet(self,sheet_name):
        if sheet_name in self.get_sheet_names():#判断新建的sheet名称是否存在
            return

        self.wb.create_sheet(sheet_name)
        self.wb.save(self.file_path)


    def get_a_line(self,row_no):#获取某一行的行对象
        rows = []#存储一下行对象
        for row in self.sheet.iter_rows():#遍历所有的行
            rows.append(row)
        if not isinstance(row_no,int):
            print("输入的行号 %s 不是一个整数" %row_no)
            return None
        if not 0<=row_no<len(rows):
            print("输入的行号 %s 超过范围" %row_no)
            return  None

        return rows[row_no]


    def get_a_line_values(self,row_no):
        values = []
        for cell in self.get_a_line(row_no):
            values.append(cell.value)
        return values


    def get_a_column(self,col_no):#获取某一列的行对象
        cols = []#存储一下列对象
        for row in self.sheet.iter_cols():#遍历所有的列
            cols.append(row)
        if not isinstance(col_no,int):
            print("输入的列号 %s 不是一个整数" %col_no)
            return None
        if not 0<=col_no<len(cols):
            print("输入的列号 %s 超过范围" %col_no)
            return  None

        return cols[col_no]


    def get_a_column_values(self,col_no):
        values = []
        for cell in self.get_a_column(col_no):
            values.append(cell.value)
        return values

    def get_cell_value(self,row_no,col_no):#获取某一个单元格的数据值
        if (not isinstance(row_no,int)) or (not isinstance(col_no,int)):
            print("输入的行号 %s 或者 列号 %s 有误" %(row_no,col_no))
            return None

        if not 0<= row_no<self.get_max_row_no():
            print("输入的行号 %s 超过最大范围")
            return None

        if not 0<= col_no<self.get_max_col_no():
            print("输入的列号 %s 超过最大范围")
            return None

        return self.sheet.cell(row = row_no + 1,column = col_no + 1).value

    def write_cell_value(self,row_no,col_no,value,colour = None):
        if (not isinstance(row_no,int)) or (not isinstance(col_no,int)):
            print("输入的行号 %s 或者 列号 %s 有误" %(row_no,col_no))
            return None
        if colour:
            if "red" in colour:
                self.sheet.cell(row = row_no + 1,column = col_no + 1).font = Font(color = colors.RED)
            elif "green" in colour:
                self.sheet.cell(row=row_no + 1, column=col_no + 1).font = Font(color=colors.GREEN)
            else:
                self.sheet.cell(row=row_no + 1, column=col_no + 1).font = Font(color=colors.BLACK)

        bd = Side(style = 'thin',color = "000000")
        self.sheet.cell(row = row_no + 1,column = col_no + 1).border = Border(left = bd,top=bd,right = bd,bottom = bd)
        self.sheet.cell(row = row_no + 1,column = col_no + 1).value = value


    def write_cell_time(self,row_no,col_no):
        current_time = TimeUtil().get_chinesedatetime()
        self.write_cell_value(row_no,col_no,current_time)



    def save(self):

        self.wb.save(self.file_path)


    def get_sheet_all_cell(self):
        data = []
        for row in self.sheet.iter_rows():
            row_cell_objects = []






if  __name__ == "__main__":
    #excel_file = ExcelUtil("d:\\a.txt")
    excel_file = ExcelUtil(r"D:\test\test_data_driven_framework0608\TestData\126邮箱联系人.xlsx")
    #print(excel_file.get_sheet_names())
    # print(excel_file.set_sheet_names_by_index(1))
    # print(excel_file.set_sheet_names_by_index(2))
    # print(excel_file.set_sheet_names_by_index(3))
    # #print(excel_file.set_sheet_names_by_index(-1))
    #print(excel_file.set_sheet_by_name("123"))
    excel_file.set_sheet_by_name("126账号")
    #print(excel_file.get_sheet_all_data())
    # print(excel_file.get_max_row_no())
    # print(excel_file.get_max_col_no())
    #data = excel_file.get_sheet_all_data()
    #excel_file.create_new_sheet("Test Result")
    # print(data)
    #excel_file.set_sheet_by_name("Test Result")
    #excel_file.write_lines_in_sheet(data)
    #print(excel_file.get_a_line('a'))
    #print(excel_file.get_a_line_values(1))
    #print(excel_file.get_a_column_values(0))
    #print(excel_file.get_cell_value(1,1))
    #print(excel_file.write_cell_value(5,0,"北京加油","red"))
    #print(excel_file.write_col_in_sheet(8,["2020-08-09","2020-04-27"]))
    #excel_file.write_cell_time(8,0)
    excel_file.write_a_line_in_sheet(["abc","123"])
    excel_file.save()